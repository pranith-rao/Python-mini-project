import speech_recognition as sr
import os
import pyttsx3
import yfinance as yf
import time
from gtts import gTTS
from datetime import datetime
import webbrowser
import random
import pandas as pd
import pymysql
import openpyxl
from openpyxl.styles import Font
from docx import Document
from docx2pdf import convert
import PyPDF2 as pdf
import os
import json
import csv

class person:
    name = ''
    def setName(self, name):
        self.name = name

#connecting to the database
db = pymysql.connect('localhost', 'root', None, 'contactbook')
cursor = db.cursor()


#dummy data
df = {       'First_name': [''],
             'Last_name': [''],
             'Mobile_no': []
             }

#insert function
def insert_data():
    speak("Enter first name")
    first_name = str(input("Enter first name : "))
    speak("Enter last name")
    last_name = str(input("Enter last name : "))
    speak("Enter mobile number")
    mobile = int(input("Enter mobile number : "))
    sql = "INSERT INTO CONTACTS(FIRST_NAME," \
          "LAST_NAME, CONTACT) " \
          "VALUES ('%s', '%s', '%s' )" % \
          (first_name, last_name, mobile)
    try:
        cursor.execute(sql)
        db.commit()
        speak("successfully inserted")
    except Exception:
        db.rollback()
    

#display function
def Dis():
    sql = """SELECT * FROM contacts"""
    cursor.execute(sql)
    rows = cursor.fetchall()
    
    print('first name     lastname     Contact no')
    for row in rows:
        print(row)


#delete function
def delete_data():
    speak("Enter the first name of the person's record to be deleted")
    var = input("Enter the first name of the person's record to be deleted : ")
    sql = "DELETE FROM CONTACTS WHERE first_name = '%s'" % var
    try:
        cursor.execute(sql)
        db.commit()
        speak("Deleted Successfully")
    except Exception:
        db.rollback()
    


#function to convert to csv
def convert_csv():
    df = pd.DataFrame()
    sql = """SELECT * FROM CONTACTS"""
    cursor.execute(sql)
    
    rows = cursor.fetchall()
    for row in rows:
        df2 = {'First_name': row[1], 'Last_name': row[2], 'Mobile_no': row[3]}
        df = df.append(df2, ignore_index=True)
    outputFile = open('Saves\phonebook.csv', 'w', newline='')
    outputWriter = csv.writer(outputFile)
    for index, row in df.iterrows(): 
        outputWriter.writerow([row["First_name"], row["Last_name"], row["Mobile_no"]])    
    speak("Successfully converted to csv.")
    


#function to convert to excel
def convert_excel():
    sql = """SELECT * FROM CONTACTS"""
    cursor.execute(sql)
    rows = cursor.fetchall()
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('Contact', 0)
    count = 2
    sheet.cell(row=1, column=1).value = 'First name'
    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=2).value = 'Last name'
    sheet.cell(row=1, column=2).font = Font(bold=True)
    sheet.cell(row=1, column=3).value = 'Contact'
    sheet.cell(row=1, column=3).font = Font(bold=True)
    for x in rows:
        sheet.cell(row=count, column=1).value = x[1]
        sheet.cell(row=count, column=2).value = x[2]
        sheet.cell(row=count, column=3).value = x[3]
        count += 1
    wb.save(r'Saves\contacts.xlsx')
    speak("Successfully converted to excel document.")



#function to convert into word doc
def convert_word(  loc = r'Saves\phone1.docx' ):
    sql = """SELECT * FROM CONTACTS"""
    cursor.execute(sql)
    records = cursor.fetchall()
    document = Document()
    table = document.add_table(rows=1, cols=3)
    hdr_cells = table.rows[0].cells
    hdr_cells[0].text = 'First_name'
    hdr_cells[1].text = 'last_name'
    hdr_cells[2].text = 'Contact'
    for i in records:
        print(i)
        row_cells = table.add_row().cells
        row_cells[0].text = i[1]
        row_cells[1].text = i[2]
        row_cells[2].text = str(i[3])
    document.save(loc)
    speak("Successfully converted to word document.")


#function to convert into Pdf
def convert_Pdf():
    convert_word('wrdoc.docx')
    socadd = r"D:\CSE\python\python project\wrdoc.docx"
    desadd = r"D:\CSE\python\python project\pdf.pdf"
    convert(socadd,desadd)
    pdfwrit = pdf.PdfFileWriter()
    rd_pdf = open('pdf.pdf','rb')
    pdf_con = pdf.PdfFileReader(rd_pdf)
    pdf_page = pdf_con.getPage(0)
    pdfwrit.addPage(pdf_page)
    con_pdf = open(r'Saves\contacts.pdf','wb')
    pdfwrit.write(con_pdf)
    con_pdf.close()
    rd_pdf.close()
    os.remove('wrdoc.docx')
    os.remove('pdf.pdf')
    speak("Successfully converted to PDF.")


#function to convert into JSON
def convert_json():
    sql = """SELECT * FROM CONTACTS"""
    cursor.execute(sql)
    records = cursor.fetchall()
    dict = {'FirstName': [], 'LastName': [], 'Contact': []}
    for row in records:
        dict['FirstName'].append(row[1])
        dict['LastName'].append(row[2])
        dict['Contact'].append(row[3])
    with open(r'saves\person.json', 'w') as json_file:
        json.dump(dict, json_file)
    speak("Successfully converted to JSON document.")



#setting up pyttsx properties
def speak(text):
    engine = pyttsx3.init()
    voice_id = "HKEY_LOCAL_MACHINE\SOFTWARE\Microsoft\Speech\Voices\Tokens\TTS_MS_EN-US_ZIRA_11.0"
    engine.setProperty('rate', 170)
    engine.setProperty('voice',voice_id)
    engine.say(text)
    engine.runAndWait()


def get_audio(ask=False):
    r = sr.Recognizer()
    with sr.Microphone() as source:
        r.adjust_for_ambient_noise(source)     #  Eliminating noise  
        print("I'm Listening....")
        audio = r.listen(source, phrase_time_limit=5)        #  Listening to Audio
        text =""
        try:
            text = r.recognize_google(audio)                 #  Recognizing Audio using the Google Speech API
            print(text)
        except Exception:
            speak("Sorry couldn't recognize your voice")
    return text

def there_exists(terms):
    for term in terms:
        if term in voice_data:
            return True

def respond(voice_data):
    # 1: greeting
    if there_exists(['hey','hi','hello']):
        greetings = [f"Hey, how can I help you {person_obj.name}", f"hey, what's up? {person_obj.name}", f"I'm listening {person_obj.name}", f"how can I help you? {person_obj.name}", f"hello {person_obj.name}"]
        greet = greetings[random.randint(0,len(greetings)-1)]
        speak(greet)

    # remember name in person object
    if there_exists(["my name is"]):
        person_name = voice_data.split("is")[-1].strip()
        speak(f"okay, i will remember that {person_name}")
        person_obj.setName(person_name) 

    # 3: greeting
    if there_exists(["how are you","how are you doing"]):
        speak(f"I'm very well, thanks for asking {person_obj.name}")

    # 4: time
    if there_exists(["what's the time","tell me the time","what time is it"]):
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        speak(current_time)

    # 5: search google
    if there_exists(["search for"]) and 'youtube' not in voice_data:
        search_term = voice_data.split("for")[-1]
        url = f"https://google.com/search?q={search_term}"
        webbrowser.get().open(url)
        speak(f'Here is what I found for {search_term} on google')

    # 6: search youtube
    if there_exists(["youtube"]):
        search_term = voice_data.split("for")[-1]
        url = f"https://www.youtube.com/results?search_query={search_term}"
        webbrowser.get().open(url)
        speak(f'Here is what I found for {search_term} on youtube')

    # explaining the program
    if there_exists(["explain","function","program","what are you","who are you"]):
        #replace the content with the program
        explain = " I am a speech recognition system and im cool, i was built by Rakshraj, shanwill, rhea and pranith, these guys got them skills; i can convert data to excel,csv,json,pdf and word. I also connect to a database. Ain't i cool? I think i deserve a ten on ten."
        speak(explain)



    if there_exists(["show","display","contacts"]):
        #display the contacts from the database
        Dis()

    if there_exists(["insert","insert records","add data","insert data"]):
        #insert data 
        insert_data()

    if there_exists(["delete","delete records","delete data"]):
        #delete data
        delete_data()

    if there_exists(["csv","convert to csv","save"]):
        #converting to csv
        convert_csv()

    if there_exists(["PDF","convert to PDF","save as PDF"]):
        #converting to pdf
        convert_Pdf()

    if there_exists(["excel","convert to excel","save as excel"]):
        #converting to excel
        convert_excel()

    if there_exists(["word","convert to word","save as word"]):
        #converting to word
        convert_word()

    if there_exists(["JSON","convert to JSON","save as JSON","javascript object notation"]):
        #converting to json
        convert_json()

    if there_exists(["CLI","time to go old school","turn on CLI mode","turn on cli mode","cli","expert mode","command line interface","hacker"]):
        speak("Turning on cli mode.....")
        #open cli interface
        os.system('phone.py')
    

    # 7: get stock price
    if there_exists(["price of"]):
        search_term = voice_data.lower().split(" of ")[-1].strip() #strip removes whitespace after/before a term in string
        stocks = {
            
            "bitcoin":"BTC-INR"
        }
        try:
            stock = stocks[search_term]
            stock = yf.Ticker(stock)
            price = stock.info["regularMarketPrice"]

            speak(f'price of {search_term} is {price} {stock.info["currency"]} {person_obj.name}')
        except Exception:
            speak('oops, something went wrong')
    if there_exists(["exit", "quit", "goodbye"]):
        speak("Terminating speech systems, Have a great day")
        exit()


        time.sleep(1)



person_obj = person()
while(1):
    voice_data = get_audio() # get the voice input
    respond(voice_data) # respond
