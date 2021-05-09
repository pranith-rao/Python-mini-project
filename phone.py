#adding required libraries
import pandas as pd
import pymysql
import openpyxl
from openpyxl.styles import Font
from docx import Document
from docx2pdf import convert
import PyPDF2 as pdf
import os
import json


#connecting to the database
db = pymysql.connect('localhost', 'root', None, 'contactbook')
cursor = db.cursor()
#dummy data
df = {'First_name': [''],
             'Last_name': [''],
             'Mobile_no': []
             }





#insert function
def insert_data():
    first_name = str(input("Enter your first name : "))
    last_name = str(input("Enter your last name : "))
    mobile = int(input("Enter your mobile number : "))
    sql = "INSERT INTO CONTACTS(FIRST_NAME," \
          "LAST_NAME, CONTACT) " \
          "VALUES ('%s', '%s', '%s' )" % \
          (first_name, last_name, mobile)
    try:
        cursor.execute(sql)
        db.commit()
        
    except Exception:
        db.rollback()


#display function
def Dis():
    sql = """SELECT * FROM contacts"""
    cursor.execute(sql)
    rows = cursor.fetchall()
    
    print('first name     lastname     Contact no')
    for row in rows:
        print(row)


#delete function
def delete_data():
    var = input("Enter the first name of the person's record to be deleted : ")
    sql = "DELETE FROM CONTACTS WHERE first_name = '%s'" % var
    try:
        cursor.execute(sql)
        db.commit()
        cursor.close()
    except Exception:
        db.rollback()
    print("Deleted Successfully")


#function to convert to csv
def convert_csv():
    #df = open(r"Saves\phonebook.csv",'wb')
    #df.close()
    #df = pd.read_csv(r'Saves\phonebook.csv')
    df = pd.DataFrame()
    sql = """SELECT * FROM CONTACTS"""
    cursor.execute(sql)
    
    rows = cursor.fetchall()
    for row in rows:
        df2 = {'First_name': row[1], 'Last_name': row[2], 'Mobile_no': row[3]}
        df = df.append(df2, ignore_index=True)
    outputFile = open('Saves\phonebook.csv', 'w', newline='')
    outputWriter = csv.writer(outputFile)
    for index, row in df.iterrows(): 
        outputWriter.writerow([row["First_name"], row["Last_name"], row["Mobile_no"]])    
    speak("Successfully converted to csv.")
    cursor.close()
    return df



#function to convert to excel
def convert_excel():
    sql = """SELECT * FROM CONTACTS"""
    cursor.execute(sql)
    rows = cursor.fetchall()
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('Contact', 0)
    count = 2
    sheet.cell(row=1, column=1).value = 'First name'
    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=2).value = 'Last name'
    sheet.cell(row=1, column=2).font = Font(bold=True)
    sheet.cell(row=1, column=3).value = 'Contact'
    sheet.cell(row=1, column=3).font = Font(bold=True)
    for x in rows:
        sheet.cell(row=count, column=1).value = x[1]
        sheet.cell(row=count, column=2).value = x[2]
        sheet.cell(row=count, column=3).value = x[3]
        count += 1
    wb.save(r'Saves\contacts.xlsx')



#function to convert into word doc
def convert_word(  loc = r'Saves\phone1.docx' ):
    sql = """SELECT * FROM CONTACTS"""
    cursor.execute(sql)
    records = cursor.fetchall()
    document = Document()
    table = document.add_table(rows=1, cols=3)
    hdr_cells = table.rows[0].cells
    hdr_cells[0].text = 'First_name'
    hdr_cells[1].text = 'last_name'
    hdr_cells[2].text = 'Contact'
    for i in records:
        print(i)
        row_cells = table.add_row().cells
        row_cells[0].text = i[1]
        row_cells[1].text = i[2]
        row_cells[2].text = str(i[3])
    document.save(loc)


#function to convert into Pdf
def convert_Pdf():
    convert_word('wrdoc.docx')
    socadd = r"D:\CSE\python\python project\wrdoc.docx"
    desadd = r"D:\CSE\python\python project\pdf.pdf"
    convert(socadd,desadd)
    pdfwrit = pdf.PdfFileWriter()
    rd_pdf = open('pdf.pdf','rb')
    pdf_con = pdf.PdfFileReader(rd_pdf)
    pdf_page = pdf_con.getPage(0)
    pdfwrit.addPage(pdf_page)
    con_pdf = open(r'Saves\contacts.pdf','wb')
    pdfwrit.write(con_pdf)
    con_pdf.close()
    rd_pdf.close()
    os.remove('wrdoc.docx')
    os.remove('pdf.pdf')


#function to convert into JSON
def convert_json():
    sql = """SELECT * FROM CONTACTS"""
    cursor.execute(sql)
    records = cursor.fetchall()
    dict = {'FirstName': [], 'LastName': [], 'Contact': []}
    for row in records:
        dict['FirstName'].append(row[1])
        dict['LastName'].append(row[2])
        dict['Contact'].append(row[3])
    with open(r'saves\person.json', 'w') as json_file:
        json.dump(dict, json_file)



#function to join all junction and run appropriate function
def switch(choice, ):
    if choice == 0:
        insert_data()
        print("Successfully inserted\n")
    elif choice == 1:
        Dis()
    elif choice == 2:
        delete_data()
    elif choice == 3:
        df = convert_csv()
        
        print("Operation Completed Successfully\n")
    elif choice == 4:
        convert_excel()
        print("Operation Completed Successfully\n")
    elif choice == 5:
        convert_word()
        print("Operation Completed Successfully\n")
    elif choice == 6:
        convert_Pdf()
        print("Operation Completed Successfully\n")
    elif choice == 7:
        convert_json()
        print("Operation Completed Successfully\n")
    elif choice == 8:
        os.system("speech.py")
    elif choice == 9:
        exit()
    else:
        print("invalid option")



while (True):
    print("----Phonebook----")
    print(
        "0. Insert New Record\n1. Display All Records\n2. Delete Record\n3. Save Records to CSV\n4. Save Records to Excel\n5.Save Records to DOC\n6.Save Records to pdf\n7.Save Records to json\n8. Speech Recognition\n9. Exit")
    choice = int(input("Enter your Choice : "))
    switch(choice)
    print("\n\n")

